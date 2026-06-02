# ==========================================================
# export.py — PDF Report + Excel Export
# Supports: batch (200 images) and single image
# ==========================================================

import io
import cv2
import datetime
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, Image as RLImage,
                                 PageBreak)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from angle import COCO17_NAMES

# ============================================================
# COLOR MAPS
# ============================================================

RISK_FILLS_XL = {
    "Negligible": PatternFill("solid", fgColor="92D050"),
    "Low":        PatternFill("solid", fgColor="FFFF00"),
    "Medium":     PatternFill("solid", fgColor="FFC000"),
    "High":       PatternFill("solid", fgColor="FF0000"),
    "Very High":  PatternFill("solid", fgColor="C00000"),
}
RISK_FONTS_XL = {
    "Negligible": Font(bold=True, color="000000"),
    "Low":        Font(bold=True, color="000000"),
    "Medium":     Font(bold=True, color="000000"),
    "High":       Font(bold=True, color="FFFFFF"),
    "Very High":  Font(bold=True, color="FFFFFF"),
}
RISK_COLORS_PDF = {
    "Negligible": colors.HexColor("#92D050"),
    "Low":        colors.HexColor("#FFFF00"),
    "Medium":     colors.HexColor("#FFC000"),
    "High":       colors.HexColor("#FF0000"),
    "Very High":  colors.HexColor("#C00000"),
}
RISK_ORDER = ["Negligible","Low","Medium","High","Very High"]

# ============================================================
# EXCEL EXPORT
# ============================================================

def build_excel(df_reba, coords_all):
    wb     = openpyxl.Workbook()
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF")

    def style_hdr(ws):
        for cell in ws[1]:
            cell.fill=HDR_FILL; cell.font=HDR_FONT
            cell.alignment=center; cell.border=border

    def auto_w(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[
                get_column_letter(col[0].column)].width = min(mx+4, 30)

    # Sheet 1: REBA Scores
    ws1 = wb.active; ws1.title = "REBA_Scores"
    ws1.append(list(df_reba.columns)); style_hdr(ws1)
    for _, row in df_reba.iterrows():
        ws1.append(list(row))
        r = ws1.max_row; rv = str(row.get("Risk",""))
        for c in range(1, len(df_reba.columns)+1):
            cell = ws1.cell(r,c)
            cell.alignment = center; cell.border = border
            if c == len(df_reba.columns):
                cell.fill = RISK_FILLS_XL.get(rv, PatternFill())
                cell.font = RISK_FONTS_XL.get(rv, Font(bold=True))
    ws1.freeze_panes = "A2"; auto_w(ws1)

    # Sheet 2: YOLO 2D Coordinates
    ws2 = wb.create_sheet("YOLO_2D_Coordinates")
    hdr2 = ["Frame_ID","Filename","TrackID"]
    for name in COCO17_NAMES:
        hdr2 += [f"{name}_pixel_x", f"{name}_pixel_y", f"{name}_conf"]
    ws2.append(hdr2); style_hdr(ws2)
    for row in coords_all:
        ws2.append(row["2d"])
        r = ws2.max_row
        for c in range(1, len(hdr2)+1):
            ws2.cell(r,c).alignment=center; ws2.cell(r,c).border=border
    ws2.freeze_panes = "A2"; auto_w(ws2)

    # Sheet 3: MiDaS 3D (if available)
    if any("3d" in row for row in coords_all):
        ws3 = wb.create_sheet("MiDaS_3D_Coordinates")
        hdr3 = ["Frame_ID","Filename","TrackID"]
        for name in COCO17_NAMES:
            hdr3 += [f"{name}_X", f"{name}_Y", f"{name}_Z"]
        ws3.append(hdr3); style_hdr(ws3)
        for row in coords_all:
            if "3d" in row:
                ws3.append(row["3d"])
                r = ws3.max_row
                for c in range(1, len(hdr3)+1):
                    ws3.cell(r,c).alignment=center; ws3.cell(r,c).border=border
        ws3.freeze_panes = "A2"; auto_w(ws3)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ============================================================
# PDF HELPERS
# ============================================================

def _hdr_style():
    return [
        ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0,0),(-1,0), colors.white),
        ("FONTNAME",  (0,0),(-1,0), "Helvetica-Bold"),
        ("ALIGN",     (0,0),(-1,-1),"CENTER"),
        ("VALIGN",    (0,0),(-1,-1),"MIDDLE"),
        ("GRID",      (0,0),(-1,-1), 0.3, colors.grey),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),
         [colors.white, colors.HexColor("#EBF3FB")]),
    ]

def _summary_table(df_reba, n_frames):
    avg_r = round(df_reba["REBA_Final"].mean(), 1)
    max_r = df_reba["REBA_Final"].max()
    n_ppl = df_reba["TrackID"].nunique()
    total = len(df_reba)
    md = [["Images/Frames","People","Detections","Avg REBA","Max REBA"],
          [str(n_frames), str(n_ppl), str(total), str(avg_r), str(max_r)]]
    mt = Table(md, colWidths=[3.6*cm]*5)
    mt.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0,0),(-1,0), colors.white),
        ("FONTNAME",  (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",  (0,0),(-1,-1), 10),
        ("ALIGN",     (0,0),(-1,-1), "CENTER"),
        ("VALIGN",    (0,0),(-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#EBF3FB")]),
        ("GRID",      (0,0),(-1,-1), 0.5, colors.grey),
    ]))
    return mt

def _risk_dist_table(df_reba):
    rc    = df_reba["Risk"].value_counts()
    total = len(df_reba)
    rd    = [["Risk Level","Count","Percentage"]]
    for r in RISK_ORDER:
        if r in rc.index:
            cnt = rc[r]
            rd.append([r, str(cnt), f"{cnt/total*100:.1f}%"])
    rt = Table(rd, colWidths=[6*cm,3*cm,3*cm])
    rts = _hdr_style() + [("FONTSIZE",(0,0),(-1,-1),10)]
    for i,r in enumerate([x for x in RISK_ORDER if x in rc.index], start=1):
        rts.append(("BACKGROUND",(0,i),(-1,i), RISK_COLORS_PDF[r]))
        if r in ("High","Very High"):
            rts.append(("TEXTCOLOR",(0,i),(-1,i), colors.white))
    rt.setStyle(TableStyle(rts))
    return rt

def _risk_legend():
    ld = [["Score","Risk Level","Required Action"],
          ["1",     "Negligible", "No action needed"],
          ["2-3",   "Low",        "May need action"],
          ["4-7",   "Medium",     "Action needed soon"],
          ["8-10",  "High",       "Investigate & change now"],
          ["11-15", "Very High",  "Implement change immediately"]]
    RC = [colors.white,
          colors.HexColor("#92D050"), colors.HexColor("#FFFF00"),
          colors.HexColor("#FFC000"), colors.HexColor("#FF0000"),
          colors.HexColor("#C00000")]
    lt = Table(ld, colWidths=[2.5*cm,4*cm,11*cm])
    lts = _hdr_style() + [("FONTSIZE",(0,0),(-1,-1),9)]
    for i in range(1,6):
        lts.append(("BACKGROUND",(0,i),(-1,i), RC[i]))
        if i>=4: lts.append(("TEXTCOLOR",(0,i),(-1,i), colors.white))
    lt.setStyle(TableStyle(lts))
    return lt

def _reba_score_table(df_reba):
    pcols = ["Frame_ID","Filename","TrackID",
             "Trunk_fwd°","Neck_fwd°","Knee_bend°",
             "UpperArm°","Elbow°",
             "Score_A","Score_B","Table_C","Activity",
             "REBA_Final","Risk"]
    pcols  = [c for c in pcols if c in df_reba.columns]
    pdf_df = df_reba[pcols].copy()
    td     = [pcols] + [
        [str(round(v,1)) if isinstance(v,float) else str(v) for v in row]
        for _, row in pdf_df.iterrows()
    ]
    tbl = Table(td, colWidths=[19*cm/len(pcols)]*len(pcols), repeatRows=1)
    ts  = _hdr_style() + [("FONTSIZE",(0,0),(-1,-1),6)]
    ri  = pcols.index("Risk") if "Risk" in pcols else -1
    if ri >= 0:
        for i,(_,row) in enumerate(pdf_df.iterrows(), start=1):
            rv = str(row.get("Risk",""))
            ts.append(("BACKGROUND",(ri,i),(ri,i),
                       RISK_COLORS_PDF.get(rv, colors.white)))
            if rv in ("High","Very High"):
                ts.append(("TEXTCOLOR",(ri,i),(ri,i), colors.white))
    tbl.setStyle(TableStyle(ts))
    return tbl

def _frame_to_rl_image(img_bgr, max_w, max_h):
    """Convert OpenCV BGR image to ReportLab Image object."""
    img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(img_rgb)
    ibuf    = io.BytesIO()
    pil_img.save(ibuf, format="PNG")
    ibuf.seek(0)
    ih, iw  = img_bgr.shape[:2]
    scale   = min(max_w/iw, max_h/ih)
    return RLImage(ibuf, width=iw*scale, height=ih*scale)

# ============================================================
# PDF REPORT — ALL IMAGES
# ============================================================

def build_pdf(df_reba, annotated_frames):
    """
    Build PDF report for ALL images (no limit).
    Layout:
      Page 1: Title + Summary + Risk distribution + Legend + Score table
      Page 2+: Annotated images 2 per row, 3 rows per page = 6 per page
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Styles
    t_style = ParagraphStyle("t", parent=styles["Title"],
        fontSize=18, spaceAfter=4,
        textColor=colors.HexColor("#1F3864"), alignment=TA_CENTER)
    s_style = ParagraphStyle("s", parent=styles["Normal"],
        fontSize=9, spaceAfter=10,
        textColor=colors.grey, alignment=TA_CENTER)
    h2 = styles["Heading2"]

    # ── PAGE 1: Summary ───────────────────────────────────
    story.append(Paragraph("REBA Assessment Report", t_style))
    story.append(Paragraph(
        f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"YOLOv11 + 2D Pose Estimation System", s_style))
    story.append(Spacer(1, 0.3*cm))

    n_frames = df_reba["Frame_ID"].nunique() if "Frame_ID" in df_reba.columns \
               else len(annotated_frames)

    story.append(_summary_table(df_reba, n_frames))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Risk Level Distribution", h2))
    story.append(_risk_dist_table(df_reba))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Risk Level Reference", h2))
    story.append(_risk_legend())
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("REBA Score Breakdown", h2))
    story.append(_reba_score_table(df_reba))
    story.append(PageBreak())

    # ── PAGE 2+: Annotated Images ────────────────────────
    # 2 images per row, 3 rows per page = 6 images per page
    n_total = len(annotated_frames)
    story.append(Paragraph(
        f"Annotated Frames — All {n_total} Images", h2))
    story.append(Spacer(1, 0.2*cm))

    IMG_W   = 8.5*cm    # width per image
    IMG_H   = 6.5*cm    # max height per image
    COL_W   = [IMG_W, IMG_W]

    for i in range(0, n_total, 2):
        row_cells = []
        row_caps  = []

        for j in range(2):
            if i+j < n_total:
                fid, img_bgr = annotated_frames[i+j]
                # Get filename if available
                fname = ""
                if "Filename" in df_reba.columns:
                    rows = df_reba[df_reba["Frame_ID"]==fid]
                    if len(rows)>0:
                        fname = str(rows.iloc[0]["Filename"])
                # Get REBA score for this frame
                reba_info = ""
                if len(df_reba[df_reba["Frame_ID"]==fid]) > 0:
                    frame_df = df_reba[df_reba["Frame_ID"]==fid]
                    max_r  = frame_df["REBA_Final"].max()
                    risk_v = frame_df.loc[frame_df["REBA_Final"].idxmax(),"Risk"]
                    reba_info = f"REBA: {max_r} | {risk_v}"

                rl_img = _frame_to_rl_image(img_bgr, IMG_W, IMG_H)
                cap    = f"Frame {fid+1}"
                if fname: cap += f"\n{fname}"
                if reba_info: cap += f"\n{reba_info}"
                row_cells.append(rl_img)
                row_caps.append(cap)
            else:
                row_cells.append("")
                row_caps.append("")

        # Image row
        img_row = Table([row_cells], colWidths=COL_W)
        img_row.setStyle(TableStyle([
            ("ALIGN",        (0,0),(-1,-1), "CENTER"),
            ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
            ("LEFTPADDING",  (0,0),(-1,-1), 4),
            ("RIGHTPADDING", (0,0),(-1,-1), 4),
            ("TOPPADDING",   (0,0),(-1,-1), 2),
            ("BOTTOMPADDING",(0,0),(-1,-1), 2),
        ]))
        story.append(img_row)

        # Caption row
        cap_style = ParagraphStyle("cap",
            parent=styles["Normal"],
            fontSize=7, alignment=TA_CENTER,
            textColor=colors.HexColor("#333333"),
            spaceAfter=4)
        cap_cells = [Paragraph(c.replace("\n","<br/>"), cap_style)
                     for c in row_caps]
        cap_row = Table([cap_cells], colWidths=COL_W)
        cap_row.setStyle(TableStyle([
            ("ALIGN",  (0,0),(-1,-1),"CENTER"),
            ("VALIGN", (0,0),(-1,-1),"TOP"),
        ]))
        story.append(cap_row)
        story.append(Spacer(1, 0.15*cm))

        # Page break every 6 images (3 rows × 2 cols)
        if (i+2) % 6 == 0 and (i+2) < n_total:
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    return buf
